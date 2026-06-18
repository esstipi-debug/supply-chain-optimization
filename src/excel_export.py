"""Excel workbook export — Vandeput (2020) results for what-if in Excel."""

from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(color="FFFFFF", bold=True)
TITLE_FONT = Font(bold=True, size=12)


def _autosize(ws) -> None:
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        width = max(len(str(cell.value or "")) for cell in col) + 2
        ws.column_dimensions[letter].width = min(width, 40)


def _write_table(ws, start_row: int, headers: list[str], rows: list[list[Any]]) -> int:
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=start_row, column=c, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    for r_idx, row in enumerate(rows, start_row + 1):
        for c_idx, val in enumerate(row, 1):
            ws.cell(row=r_idx, column=c_idx, value=val)
    return start_row + len(rows) + 2


def write_analysis_workbook(
    path: Path | str,
    *,
    product_id: str,
    parameters: dict[str, Any],
    results: dict[str, Any],
    gsm: dict[str, Any] | None = None,
    simulation: dict[str, Any] | None = None,
    newsvendor: dict[str, Any] | None = None,
) -> Path:
    """Write multi-sheet .xlsx workbook from analysis context."""
    out = Path(path)
    out.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws_sum = wb.active
    ws_sum.title = "Summary"

    ws_sum["A1"] = "Inventory Optimization — Vandeput (2020)"
    ws_sum["A1"].font = TITLE_FONT
    ws_sum["A2"] = f"Product: {product_id}"

    summary_rows = [[k, v] for k, v in results.items()]
    _write_table(ws_sum, 4, ["Metric", "Value"], summary_rows)
    _autosize(ws_sum)

    ws_in = wb.create_sheet("Parameters")
    ws_in["A1"] = "Inputs"
    ws_in["A1"].font = TITLE_FONT
    _write_table(ws_in, 3, ["Parameter", "Value"], [[k, v] for k, v in parameters.items()])
    _autosize(ws_in)

    ws_ref = wb.create_sheet("Formulas")
    ws_ref["A1"] = "Reference formulas (book)"
    ws_ref["A1"].font = TITLE_FONT
    formulas = [
        ("EOQ", "Q* = SQRT(2*D*k/h)"),
        ("EOQ cost", "C* = SQRT(2*D*k*h)"),
        ("Safety stock", "Ss = NORM.S.INV(alpha)*sigma*SQRT(tau)"),
        ("Fill rate", "beta = 1 - U_s / mu_x"),
        ("Optimal alpha (R,S)", "alpha* = 1 - h*R/b"),
        ("Newsvendor CR", "cu/(cu+co)"),
        ("GSM", "Ss_i = z*sigma_d*SQRT(x_tau_i)"),
    ]
    _write_table(ws_ref, 3, ["Model", "Formula"], formulas)
    _autosize(ws_ref)

    if gsm:
        ws_gsm = wb.create_sheet("GSM")
        ws_gsm["A1"] = "Multi-echelon GSM"
        ws_gsm["A1"].font = TITLE_FONT
        nodes = gsm.get("nodes", [])
        node_rows = [
            [
                n.get("index"),
                n.get("lead_time"),
                n.get("risk_period"),
                n.get("safety_stock"),
                n.get("order_up_to"),
                n.get("holding_cost"),
            ]
            for n in nodes
        ]
        next_row = _write_table(
            ws_gsm,
            3,
            ["Node", "Lead time", "x_tau", "Ss", "Local S", "h"],
            node_rows,
        )
        ws_gsm.cell(row=next_row, column=1, value="Total holding cost")
        ws_gsm.cell(row=next_row, column=2, value=gsm.get("total_holding_cost"))
        ws_gsm.cell(row=next_row + 1, column=1, value="Echelon S levels")
        ws_gsm.cell(row=next_row + 1, column=2, value=str(gsm.get("echelon_order_up_to")))
        _autosize(ws_gsm)

    if simulation:
        ws_sim = wb.create_sheet("Simulation")
        ws_sim["A1"] = "Simulation results"
        ws_sim["A1"].font = TITLE_FONT
        _write_table(ws_sim, 3, ["Metric", "Value"], [[k, v] for k, v in simulation.items()])
        _autosize(ws_sim)

    if newsvendor:
        ws_nv = wb.create_sheet("Newsvendor")
        ws_nv["A1"] = "Newsvendor (muffins example)"
        ws_nv["A1"].font = TITLE_FONT
        _write_table(ws_nv, 3, ["Metric", "Value"], [[k, v] for k, v in newsvendor.items()])
        _autosize(ws_nv)

    wb.save(out)
    return out


def gsm_allocation_to_dict(allocation) -> dict[str, Any]:
    """Serialize GSMAllocation for Excel export."""
    return {
        "risk_periods": allocation.risk_periods,
        "total_holding_cost": allocation.total_holding_cost,
        "echelon_order_up_to": allocation.echelon_order_up_to,
        "nodes": [
            {
                "index": n.index,
                "lead_time": n.lead_time,
                "risk_period": n.risk_period,
                "safety_stock": n.safety_stock,
                "order_up_to": n.order_up_to,
                "holding_cost": n.holding_cost,
            }
            for n in allocation.nodes
        ],
    }
