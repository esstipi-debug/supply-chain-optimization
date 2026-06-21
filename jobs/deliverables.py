"""Client deliverables — turn a JobReport into an Excel workbook + a report.

These are what the client receives: a recommendations workbook for action and a
written report with findings, methodology and assumptions for credibility.
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from src.export import write_summary_csv

from .inventory_optimization import JobReport, SkuRecommendation
from .pricing import PricingRec, PricingReport

STATUS_LABEL = {"ok": "On track", "high_bias": "High bias — review", "review": "Intermittent — review"}
ACTION_LABEL = {
    "raise": "Raise price",
    "lower": "Lower price",
    "hold": "Hold price",
    "inelastic": "Inelastic — test higher",
    "insufficient_data": "Insufficient price variation",
}


def _row(r: SkuRecommendation) -> dict:
    return {
        "product_id": r.product_id,
        "method": r.method,
        "policy": r.policy_kind,
        "forecast_per_period": round(r.forecast, 2),
        "order_qty_Q": round(r.order_quantity, 2) if r.order_quantity is not None else "",
        "order_up_to_S": round(r.order_up_to, 2) if r.order_up_to is not None else "",
        "reorder_point_s": round(r.reorder_point, 2),
        "safety_stock": round(r.safety_stock, 2),
        "sigma_e": round(r.error_std, 2),
        "bias": round(r.bias, 2),
        "mae": round(r.mae, 2),
        "unit_cost": round(r.unit_cost, 2),
        "inventory_value": round(r.investment, 2),
        "status": STATUS_LABEL[r.status],
    }


def write_csv(report: JobReport, path: str | Path) -> Path:
    """One row per SKU — the machine-readable deliverable."""
    return write_summary_csv([_row(r) for r in report.recommendations], path)


def write_excel(report: JobReport, path: str | Path) -> Path:
    """A two-sheet workbook: Recommendations + Summary & Assumptions."""
    out = Path(path)
    out.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()

    ws = wb.active
    ws.title = "Recommendations"
    rows = [_row(r) for r in report.recommendations]
    headers = list(rows[0].keys()) if rows else ["product_id"]
    header_fill = PatternFill("solid", fgColor="1F2A44")
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h.replace("_", " "))
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
    for ri, row in enumerate(rows, 2):
        for ci, h in enumerate(headers, 1):
            ws.cell(row=ri, column=ci, value=row[h])
    for ci, h in enumerate(headers, 1):
        width = max(len(h), *(len(str(row[h])) for row in rows)) if rows else len(h)
        ws.column_dimensions[get_column_letter(ci)].width = min(max(width + 2, 10), 32)
    ws.freeze_panes = "A2"

    s = wb.create_sheet("Summary & Assumptions")
    p = report.params
    summary = [
        ("Inventory Optimization — Summary", ""),
        ("SKUs analyzed", report.n_skus),
        ("SKUs flagged (high bias)", report.n_at_risk),
        ("Intermittent SKUs (Croston)", report.n_intermittent),
        ("Requested investment", round(report.requested_investment, 2)),
        ("Cycle-stock floor", round(report.cycle_floor, 2)),
        ("Final investment", round(report.final_investment, 2)),
        ("Budget cap", report.budget if report.budget is not None else "none"),
        ("Safety-stock scale applied", f"{report.safety_stock_scale * 100:.0f}%"),
        ("Plan feasible", "yes" if report.feasible else "no"),
        ("", ""),
        ("Assumptions", ""),
        ("Cycle service level", f"{p['service_level'] * 100:.1f}%"),
        ("Holding cost rate (of unit cost / yr)", f"{p['holding_rate'] * 100:.0f}%"),
        ("Order cost K", p["order_cost"]),
        ("Periods per year", p["periods_per_year"]),
    ]
    for ri, (k, v) in enumerate(summary, 1):
        kc = s.cell(row=ri, column=1, value=k)
        if v == "" and k:
            kc.font = Font(bold=True)
        s.cell(row=ri, column=2, value=v)
    s.column_dimensions["A"].width = 40
    s.column_dimensions["B"].width = 22

    wb.save(out)
    return out


def _fmt(n: float) -> str:
    return f"{n:,.0f}"


def write_report_md(report: JobReport, path: str | Path, *, client: str = "Client") -> Path:
    """A written report: summary, recommendations, findings, methodology."""
    out = Path(path)
    out.parent.mkdir(parents=True, exist_ok=True)
    p = report.params
    over = report.budget is not None and report.requested_investment > report.budget

    lines: list[str] = []
    lines.append(f"# Inventory Optimization — {client}\n")
    lines.append("## Executive summary\n")
    lines.append(
        f"Analyzed **{report.n_skus} SKUs**. Recommended inventory investment is "
        f"**${_fmt(report.final_investment)}**"
        + (f" against a budget of **${_fmt(report.budget)}**" if report.budget is not None else "")
        + f", at a **{p['service_level'] * 100:.0f}%** cycle service level. "
        + (f"{report.n_at_risk} SKU(s) show high forecast bias and {report.n_intermittent} are intermittent (review recommended)."
           if (report.n_at_risk or report.n_intermittent) else "All SKUs are on track.")
    )
    lines.append("")
    if report.budget is not None:
        if not report.feasible:
            lines.append(f"> ⚠️ **Infeasible at this budget.** The cycle-stock floor alone is ${_fmt(report.cycle_floor)}; raise the cap above it to fund any safety stock.\n")
        elif report.safety_stock_scale < 1:
            lines.append(f"> Safety stock scaled to **{report.safety_stock_scale * 100:.0f}%** to fit the budget. Raise the cap to ${_fmt(report.requested_investment)} to fund it fully.\n")
        elif over:
            lines.append("")

    lines.append("## Recommended policy per SKU\n")
    lines.append("| SKU | Method | Policy | Forecast/period | Order qty (Q*) | Order-up-to (S) | Reorder point (s) | Safety stock | Inv. value | Status |")
    lines.append("|---|---|---|---:|---:|---:|---:|---:|---:|---|")
    for r in report.recommendations:
        q = f"{r.order_quantity:,.0f}" if r.order_quantity is not None else "—"
        s_up = f"{r.order_up_to:,.0f}" if r.order_up_to is not None else "—"
        lines.append(
            f"| {r.product_id} | {r.method} | {r.policy_kind} | {r.forecast:,.1f} | {q} | {s_up} | "
            f"{r.reorder_point:,.0f} | {r.safety_stock:,.1f} | ${_fmt(r.investment)} | {STATUS_LABEL[r.status]} |"
        )
    lines.append("")

    flagged = [r for r in report.recommendations if r.status != "ok"]
    if flagged:
        lines.append("## Findings & flags\n")
        for r in flagged:
            if r.status == "high_bias":
                lines.append(f"- **{r.product_id}** — forecast bias {r.bias:+.1f} (|bias| ≥ 2). The forecast is consistently off; review the demand history or method before trusting the policy.")
            else:
                lines.append(f"- **{r.product_id}** — intermittent demand, forecast via Croston. Lumpy demand makes a periodic (R,S) review more robust than a fixed reorder point.")
        lines.append("")

    lines.append("## Methodology\n")
    lines.append("- **Forecast:** per-SKU demand history is forecast with simple exponential smoothing (or Croston for intermittent demand), exposing the forecast-error standard deviation σₑ used for safety stock (Vandeput 2021, §4.2.5).")
    lines.append("- **Safety stock:** `SS = z · σₑ · √L`, with `z` from the target cycle service level.")
    lines.append("- **Order quantity:** Economic Order Quantity `Q* = √(2·D·K/H)` for continuous-review (s,Q); periodic (R,S) order-up-to `S = μ·(L+R) + SS` for intermittent SKUs.")
    lines.append("- **Reorder point:** `s = μ·L + SS`.")
    lines.append("- **Budget:** when a cap is set, safety stock is scaled across the portfolio to fit while preserving cycle-stock economics.")
    lines.append("- Models from Vandeput (2020), *Inventory Optimization: Models and Simulations*.\n")

    lines.append("## Assumptions\n")
    lines.append(f"- Cycle service level: **{p['service_level'] * 100:.1f}%**")
    lines.append(f"- Holding cost: **{p['holding_rate'] * 100:.0f}%** of unit cost per year")
    lines.append(f"- Fixed order cost (K): **${_fmt(p['order_cost'])}**")
    lines.append(f"- Demand bucketed into **{int(p['periods_per_year'])} periods/year**; lead time taken from the data (or a default where absent).")
    lines.append("")
    lines.append("_Generated from the client's demand data. Figures are decision support; validate cost and lead-time inputs against your systems before ordering._")

    out.write_text("\n".join(lines), encoding="utf-8")
    return out


def write_all(report: JobReport, out_dir: str | Path, *, client: str = "Client") -> dict[str, Path]:
    """Write the full deliverable set into out_dir."""
    d = Path(out_dir)
    return {
        "excel": write_excel(report, d / "inventory_plan.xlsx"),
        "report": write_report_md(report, d / "report.md", client=client),
        "csv": write_csv(report, d / "summary.csv"),
    }


# ---- pricing deliverables ----------------------------------------------------

def _price_row(r: PricingRec) -> dict:
    return {
        "product_id": r.product_id,
        "current_price": round(r.current_price, 2),
        "optimal_price": round(r.optimal_price, 2) if r.optimal_price is not None else "",
        "unit_cost": round(r.unit_cost, 2),
        "elasticity": round(r.elasticity, 2),
        "r_squared": round(r.r_squared, 2),
        "obs": r.n_points,
        "demand_change_pct": round(r.demand_change_pct, 1) if r.demand_change_pct is not None else "",
        "profit_uplift_pct": round(r.profit_uplift_pct, 1) if r.profit_uplift_pct is not None else "",
        "action": ACTION_LABEL[r.action],
        "confident": "yes" if r.confident else "no",
    }


def write_pricing_csv(report: PricingReport, path: str | Path) -> Path:
    return write_summary_csv([_price_row(r) for r in report.recommendations], path)


def write_pricing_report_md(report: PricingReport, path: str | Path, *, client: str = "Client") -> Path:
    """Written price-optimization report: summary, per-SKU table, methodology."""
    out = Path(path)
    out.parent.mkdir(parents=True, exist_ok=True)

    lines: list[str] = []
    lines.append(f"# Price Optimization — {client}\n")
    lines.append("## Executive summary\n")
    lines.append(
        f"Analyzed **{report.n_skus} SKUs**. **{report.n_actionable}** have a confident price move "
        f"(raise/lower); {report.n_inelastic} are inelastic and {report.n_insufficient} lack enough "
        "price variation to estimate. Recommendations maximize unit margin under a constant-elasticity "
        "demand model fitted to each SKU's price/quantity history."
    )
    lines.append("")
    lines.append("## Recommended price per SKU\n")
    lines.append("| SKU | Current | Optimal | Unit cost | Elasticity | R² | Obs | Δ demand | Profit uplift | Action | Confident |")
    lines.append("|---|---:|---:|---:|---:|---:|---:|---:|---:|---|---|")
    for r in report.recommendations:
        opt = f"${r.optimal_price:,.2f}" if r.optimal_price is not None else "—"
        dchg = f"{r.demand_change_pct:+.0f}%" if r.demand_change_pct is not None else "—"
        upl = f"{r.profit_uplift_pct:+.0f}%" if r.profit_uplift_pct is not None else "—"
        lines.append(
            f"| {r.product_id} | ${r.current_price:,.2f} | {opt} | ${r.unit_cost:,.2f} | "
            f"{r.elasticity:.2f} | {r.r_squared:.2f} | {r.n_points} | {dchg} | {upl} | "
            f"{ACTION_LABEL[r.action]} | {'yes' if r.confident else 'no'} |"
        )
    lines.append("")

    lines.append("## Methodology\n")
    lines.append("- **Elasticity:** per-SKU log-log regression of quantity on price (ε = slope of ln q vs ln p), with R² as a fit-quality signal.")
    lines.append("- **Optimal price:** constant-elasticity profit maximum `p* = c · ε/(ε+1)`, valid when demand is elastic (ε < −1). Inelastic SKUs (ε ≥ −1) have no interior optimum — test a higher price.")
    lines.append("- **Profit uplift / demand change:** modeled against the fitted curve relative to the current (median) price.")
    lines.append("- **Confidence:** flagged only when R² ≥ 0.5, ≥ 4 price observations, and the move is within a sane range of the current price.\n")
    lines.append("## Assumptions & caveats\n")
    cr = report.params["cost_ratio"]
    lines.append(f"- Unit cost {'taken from the data' if report.params['has_cost_column'] else f'assumed at {cr * 100:.0f}% of current price'} (no cost column → margin is an estimate).")
    lines.append("- A single-product, constant-elasticity model: it ignores cross-product effects, competitor moves, and capacity. Validate before repricing, ideally with a live price test.")
    lines.append("")
    lines.append("_Decision support generated from the client's price/quantity history._")

    out.write_text("\n".join(lines), encoding="utf-8")
    return out


def write_pricing_excel(report: PricingReport, path: str | Path) -> Path:
    out = Path(path)
    out.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "Price recommendations"
    rows = [_price_row(r) for r in report.recommendations]
    headers = list(rows[0].keys()) if rows else ["product_id"]
    fill = PatternFill("solid", fgColor="1F2A44")
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h.replace("_", " "))
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = fill
    for ri, row in enumerate(rows, 2):
        for ci, h in enumerate(headers, 1):
            ws.cell(row=ri, column=ci, value=row[h])
    for ci, h in enumerate(headers, 1):
        width = max(len(h), *(len(str(row[h])) for row in rows)) if rows else len(h)
        ws.column_dimensions[get_column_letter(ci)].width = min(max(width + 2, 10), 28)
    ws.freeze_panes = "A2"
    wb.save(out)
    return out


def write_pricing_all(report: PricingReport, out_dir: str | Path, *, client: str = "Client") -> dict[str, Path]:
    d = Path(out_dir)
    return {
        "excel": write_pricing_excel(report, d / "price_recommendations.xlsx"),
        "report": write_pricing_report_md(report, d / "pricing_report.md", client=client),
        "csv": write_pricing_csv(report, d / "pricing_summary.csv"),
    }
