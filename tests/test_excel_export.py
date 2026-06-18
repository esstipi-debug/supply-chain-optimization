"""Tests for Excel workbook export."""

from pathlib import Path

import pytest
from openpyxl import load_workbook

from src.excel_export import write_analysis_workbook


def test_write_analysis_workbook_sheets(tmp_path: Path):
    path = tmp_path / "test.xlsx"
    write_analysis_workbook(
        path,
        product_id="SKU-A",
        parameters={"D": 1000, "h": 1.75},
        results={"Q*": 239},
        gsm={"nodes": [], "total_holding_cost": 485, "echelon_order_up_to": (1, 2, 3)},
        simulation={"fill_rate": 0.95},
        newsvendor={"Q*": 4},
    )
    wb = load_workbook(path)
    assert "Summary" in wb.sheetnames
    assert "Parameters" in wb.sheetnames
    assert "Formulas" in wb.sheetnames
    assert "GSM" in wb.sheetnames
    assert wb["Summary"]["B5"].value == 239
